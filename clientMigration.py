# Author: BornSmasher
#
# This program imports Pet and Client data from 123Pet and transforms it into a single
# Customer Object recognized by Revelation Pet software.
#
# inputs:    123Pet Client Excel File (.xlsx)
#            123Pet Pet Excel File (.xlsx)
#            123Pet Vaccination Excel File (.xlsx)
#
# OutPuts:   RevelationPet Customer Excel File (.xlsx)
#
###############################################################################################


from Customer import Customer
import openpyxl
import settings


# Import Excel Files for 123pet and RevelationPet
wb_123pets = openpyxl.load_workbook(settings.pet123_filepath)
wb_123vaccines = openpyxl.load_workbook(settings.vaccine123_filepath)
wb_123clients = openpyxl.load_workbook(settings.client123_filepath)
wb_revclients = openpyxl.load_workbook(settings.clientrev_filepath)


# Create 123Pet & RevPet Objects from Excel Files
def import_object(excel_filepath):
    data_object = {}
    # Get Sheet Name
    wb = openpyxl.load_workbook(excel_filepath)
    sheet_names = wb.get_sheet_names()
    ws = wb[sheet_names[0]]
    first_row = ws[1:ws.max_column][0]
    for cell in first_row:
        data_object[cell.value] = None
    return data_object


# Import 123Pet and RevPet Data Into a list of Objects
def import_data(excel_filepath):
    object_list = []
    wb = openpyxl.load_workbook(excel_filepath)
    sheet_names = wb.get_sheet_names()
    ws = wb[sheet_names[0]]
    header = [cell.value for cell in ws[1]]
    for row in ws.rows:
        values = {}
        for key, cell in zip(header, row):
            values[str(key)] = str(cell.value)
        object_list.append(values)
    object_list.pop(0)
    return object_list


# Get Excel File Header
def get_header(excel_filepath):
    wb = openpyxl.load_workbook(excel_filepath)
    sheet_names = wb.get_sheet_names()
    ws = wb[sheet_names[0]]
    header = [cell.value for cell in ws[1]]
    return header


# Add Header to Worksheet
def add_header(header, row, ws):
    for col, val in enumerate(header, start=1):
        ws.cell(row=row, column=col).value = val


# Find 123 Pets for a 123 Client
def find_pets123(client_id, pet123_list):
    pets = []
    for pet in pet123_list:
        if str(pet["ClientID"]) == str(client_id):
            pets.append(pet)
    return pets


# Find 123 Vaccines for a 123 Pet
def find_vaccines123(pets, vaccine123_list):
    vaccines = {}
    for pet in pets:
        pet_id = int(pet["PetID"])
        for vaccine in vaccine123_list:
            if int(pet_id) == int(vaccine["Pet ID"]):
                vaccine_name = vaccine["Vaccination"].lower()
                expiration_date = vaccine["Expiration Date"]
                if pet_id in vaccines:
                    vaccines[pet_id][vaccine_name] = expiration_date
                else:
                    vaccines[pet_id] = {vaccine_name: expiration_date}
    return vaccines


# Construct Client Rows
def build_client_rows(client_id, pet123_list):
    rows = []
    pets = find_pets123(client_id, pet123_list)


# Write RevelationPet Customer to Excel File
def write_clientrev(client123, pet123):
    # Write Header to file
    revpet_header = get_header(settings.clientrev_filepath)
    wb = openpyxl.Workbook()
    ws = wb.active
    # Write Header to Worksheet
    for col, val in enumerate(revpet_header, start=1):
        ws.cell(row=1, column=col).value = val

    # Write Customers to Workshhet
    for customer in customers:
        write_customer(customer)
    # Write to File
    wb.save(settings.new_clientrev_filepath)


def main():
    debug = False
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    pet123_list, vaccine123_list, client123_list, clientrev_list = [], [], [], []

    # Introduction
    print("Importing 123Pet Data....")
    print("")

    # Import List of 123Pet Pet Objects
    pet123_list = import_data(settings.pet123_filepath)
    print("123Pet Data has " + str(len(pet123_list)) + " Pets")

    # Import List of 123Pet Client Objects
    client123_list = import_data(settings.client123_filepath)
    print("123Pet Data has " + str(len(client123_list))) + " Clients"

    # Import List of 123Pet Vaccine Objects
    vaccine123_list = import_data(settings.vaccine123_filepath)
    print("123Pet Data has " + str(len(vaccine123_list))) + " Vaccinations"

    # Import Revelation Pets Client Object
    clientrev = import_object(settings.clientrev_filepath)

    # Create Revelation Pets Client List with Empty Client Objects
    for x in range(0, len(client123_list)):
        clientrev_list.append(clientrev.copy())

    print("RevPet Data has " + str(len(clientrev_list))) + " Clients"

    # Load RevelationPet Client Objects with 123Pet Client and Pet Data
    print("Loading 123Pet Data to RevelationPet Objects:")
    cust_list = []
    for client in client123_list:
        client_id = client["ClientID"]
        pets = find_pets123(client_id,
                            pet123_list)
        vaccines = find_vaccines123(pets,
                                    vaccine123_list)
        cust_list.append(Customer(client,
                                  pets,
                                  vaccines))

    ###################################
    # Build Revelation Pet Excel File #
    ###################################

    #### Add Header to Worksheet ####
    header = get_header(settings.clientrev_filepath)
    row = 1
    add_header(header,
               row,
               ws)

    #### Add Customers and Pets to Worksheet####
    row += 1
    for cust in cust_list:
        # input = raw_input("Hit enter to print a customer")
        # pprint(cust.__dict__)
        # for pet in cust.Pets:
        #     pprint(pet.__dict__)
        cust.print2ws(ws, row)
        if not cust.Pets:
            row += 1
        else:
            for pet in cust.Pets:
                pet.print2ws(ws, row)
                row += 1

    #### Print To Excel File ####
    wb.save(settings.new_clientrev_filepath)


if __name__ == "__main__":
    main()